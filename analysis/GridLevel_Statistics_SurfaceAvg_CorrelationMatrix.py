# =============================================================================
# Imports 
# =============================================================================
import pandas as pd
import numpy as np
from pathlib import Path
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import kendalltau
import warnings
warnings.filterwarnings('ignore')

# Set global font properties
plt.rcParams['font.size'] = 12
plt.rcParams['font.family'] = 'Times New Roman'
plt.rcParams['font.style'] = 'normal'
# =============================================================================


# =============================================================================
# Functions
# =============================================================================
def kendall_corr_with_pvalues(df):
    cols = df.columns
    kendall_corr = pd.DataFrame(index=cols, columns=cols, dtype=float)
    p_values = pd.DataFrame(index=cols, columns=cols, dtype=float)
    
    for col1 in cols:
        for col2 in cols:
            if col1 == col2:
                kendall_corr.loc[col1, col2] = 1.0
                p_values.loc[col1, col2] = 0.0
            else:
                sub = df[[col1,col2]].dropna()
                # sub = df.dropna(subset=[col1,col2])
                corr, p_value = kendalltau(sub[col1], sub[col2])
                kendall_corr.loc[col1, col2] = corr
                p_values.loc[col1, col2] = p_value
    
    return kendall_corr, p_values

def remove_bad_years(df):
    parambadyears = {
        'MLD': [1995, 2011],
        'QI': [1995],
        'max_N2': [1995, 2011],
        
        'WWUpper': [None],
        'WWLower': [None],
        'WWThickness': [None],
        'WW%Obs': [None],
        'WWMinTemp': [None],
        'WWMinTempDepth': [None],
        'SIAdvance': [None],
        'SIRetreat': [None],
        'SIDuration': [None],
        'IceDays': [None],
        'SIRetrProx': [None],
        'SIExtent': [None],
        'SIArea': [None],
        'OWArea': [None],
        'TotalSIConc': [None],
        
        'WWMedTemp': [None],
        'WWMedSal': [None],
        'WWMedDens': [None],
        
        'BttmTemp': [1995],
        'BttmSal': [1995],
        'BttmDens': [1995, 2017, 2018],
        
        'Temperature': [None], 
        'Salinity': [None],
        'Density': [1995], 
        
        # 'Temp_ML': [1995], 
        # 'Sal_ML': [1995],
        # 'Dens_ML': [1995],
        
        'MinTemp': [1995],
        'MinTempDepth': [1995],
        
        'TCaro:Chla': [None], 
        'PSC:Chla': [None],
        'PPC:Chla': [None], 
        'PrimPPC:Chla': [None],
        'SecPPC:Chla': [None],
        
        'PPC:TCaro': [None],
        'PSC:TCaro': [None],
        
        # 'TCaro:TAcc': [None],
        # 'PSC:TAcc': [None],
        # 'PPC:TAcc': [None],
        
        # 'Allo:PPC': [None],
        # 'Diadino:PPC': [None],
        # 'Diato:PPC': [None],
        # 'DD+DT:PPC': [None],
        # 'Zea:PPC': [None],
        # 'BCar:PPC': [None],
        # 'PrimPPC:PPC': [None],
        # 'SecPPC:PPC': [None],
        
        # 'Fuco:PSC': [None],
        # 'Hex-Fuco:PSC': [None],
        # 'But-Fuco:PSC': [None],
        # 'Perid:PSC': [None],
        
        # 'Allo:TCaro': [None],
        # 'Diadino:TCaro': [None],
        # 'Diato:TCaro': [None],
        # 'DD+DT:TCaro': [None],
        # 'Zea:TCaro': [None],
        # 'BCar:TCaro': [None],
        # 'Fuco:TCaro': [None],
        # 'Hex-Fuco:TCaro': [None],
        # 'But-Fuco:TCaro': [None],
        # 'Perid:TCaro': [None],
        
        # 'Allo': [None],
        # 'Diadino': [None],
        # 'Diato': [None],
        # 'DD+DT': [None],
        # 'Zea': [None],
        # 'BCar': [None],
        # 'Fuco': [None],
        # 'Hex-Fuco': [None],
        # 'But-Fuco': [None],
        # 'Perid': [None],
        
        # 'mPF': [None],
        # 'nPF': [None],
        # 'pPF': [None],
        
        'Chlorophylla': [None],
        # 'POC': [1994, 2009, 2015, 2016],
        'PrimaryProduction': [2019],
        'SpecPrimProd': [None],
        
        'Diatoms': [None],
        'Cryptophytes': [None],
        'MixedFlagellates': [None],
        'Type4Haptophytes': [None],
        'Prasinophytes': [None],
        
        # 'DiatomBiomass': [1993, 1994, 1998, 2015],
        # 'CryptophyteBiomass': [1993, 1994],
        # 'MixedFlagellateBiomass': [2013, 2014, 2015, 2016],
        # 'Type4HaptophyteBiomass': [1993, 1994, 1998, 2009],
        # 'PrasinophyteBiomass': [1993, 1998, 2013, 2017],
        
        # 'TAcc2:POC': [2009, 2015, 2016],
        
        # 'Chla:POC': [1994, 2009, 2015, 2016],
        # 'TCaro:POC': [2009, 2015, 2016],
        # 'TAcc:POC': [2009, 2015, 2016],
        # 'TPig:POC': [2009, 2015, 2016],
        # 'TAcc:Chla': [None],
        
        'SiO4': [1998],
        'PO4': [1998],
        'NO2': [1998],
        'NO3': [1998],
        'NO3plusNO2': [None],
        
        # 'FIRERho': [2015], #these have not been adjusted/blank corrected/filtered
        # 'FIRESigma': [2015], #these have not been adjusted/blank corrected/filtered
        # 'FIRE_FvFm': [2015], #these have not been adjusted/blank corrected/filtered
        
        # 'TCaro:POC': [1994],
        
        'Evenness': [None]}
    
    for param, bad_years in parambadyears.items():
        if param in df.columns and bad_years is not None:
            df.loc[df['Year'].isin(bad_years), param] = np.nan
    
    return df

def fill_missing_years(df):    
    all_years = np.arange(1991, 2021)
    existing_years = df['Year'].unique()
    missing_years = np.setdiff1d(all_years, existing_years)
    missing_data = pd.DataFrame({'Year': missing_years})
    missing_data = missing_data.assign(**{column: np.nan for column in df.columns if column != 'Year'})
    df = pd.concat([df, missing_data], ignore_index=True)
    df = df.sort_values('Year')
    return df
# =============================================================================


# =============================================================================
# Load Core Dataframe
# =============================================================================
current_directory = Path.cwd()
absolute_path = Path("local data/")
filename = Path("PALLTER_StationLevel_SurfaceAvgCoreDataframe.csv")
loadpath = str(current_directory / absolute_path / filename)
df_SM = pd.read_csv(loadpath)

#calculate yearly medians
temp = remove_bad_years(df_SM)
df_YM = temp.groupby(['Year']).median().reset_index()
df_YM = fill_missing_years(df_YM)

#determine climatological stats
climatology_stats = df_YM.describe()
# =============================================================================


# =============================================================================
# Plot Correlation Matrix Heatmap
# =============================================================================
#define list of parameters to compare
paramlist = [
    'MLD',
    'QI',
    'max_N2',
    
    'WWUpper',
    'WWLower',
    'WWThickness',
    'WW%Obs',
    
    # 'WWMinTemp',
    # 'WWMinTempDepth',
    # 'MinTemp',
    # 'MinTempDepth',
    
    'Temperature', 
    'Salinity',
    'Density',
    
    'WWMedTemp',
    'WWMedSal',
    'WWMedDens',
    
    # 'BttmTemp',
    # 'BttmSal',
    # 'BttmDens',
    
    # 'SIAdvance',
    # 'SIRetreat',
    'SIDuration',
    # 'IceDays',
    'SIRetrProx',
    # 'SIExtent',
    # 'SIArea',
    # 'OWArea',
    # 'TotalSIConc', 
    
    'TCaro:Chla', 
    'PSC:Chla',
    'PPC:Chla',
    
    'PrimPPC:Chla',
    'SecPPC:Chla',
    
    'PPC:PSC',
    
    'PPC:TCaro',
    'PSC:TCaro',
    
    # 'Allo:PPC',
    # 'Diadino:PPC',
    # 'Diato:PPC',
    # 'DD+DT:PPC',
    # 'Zea:PPC',
    # 'BCar:PPC',
    'PrimPPC:PPC',
    'SecPPC:PPC',
    
    # 'Fuco:PSC',
    # 'Hex-Fuco:PSC',
    # 'But-Fuco:PSC',
    # 'Perid:PSC',
    
    # 'Allo:TCaro',
    # 'Diadino:TCaro',
    # 'Diato:TCaro',
    # 'DD+DT:TCaro',
    # 'Zea:TCaro',
    # 'BCar:TCaro',
    # 'Fuco:TCaro',
    # 'Hex-Fuco:TCaro',
    # 'But-Fuco:TCaro',
    # 'Perid:TCaro',
    
    # 'Allo',
    # 'Diadino',
    # 'Diato',
    # 'Zea',
    # 'BCar',
    # 'Fuco',
    # 'Hex-Fuco',
    # 'But-Fuco',
    # 'Perid',
    
    # 'mPF',
    # 'nPF',
    # 'pPF',
    
    'Chlorophylla',
    'PrimaryProduction',
    'SpecPrimProd',
    
    'Diatoms',
    'Cryptophytes',
    'MixedFlagellates',
    'Type4Haptophytes',
    'Prasinophytes',
    
    # 'DiatomBiomass',
    # 'CryptophyteBiomass',
    # 'MixedFlagellateBiomass',
    # 'Type4HaptophyteBiomass',
    # 'PrasinophyteBiomass',
    
    # 'TAcc2:POC',
    
    # 'Chla:POC',
    # 'TCaro:POC',
    # 'TAcc:POC',
    # 'TPig:POC',
    # 'TAcc:Chla',
    
    # 'SiO4',
    # 'PO4',
    # 'NO2',
    # 'NO3',
    # 'NO3plusNO2',
    # 'POC',
    
    'Evenness'
    ]

#define matrix
Matrix = df_YM[paramlist].reset_index(drop=True)

#calculate kendall correlation and p-values
corr_matrix, p_value_matrix = kendall_corr_with_pvalues(Matrix)

#create a mask for the duplicate upper triangle values of the matrix
mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=1)

#plot heatmap of kendal_tau correlations w/ p-values as numbers 
fig = plt.figure(figsize=(20, 20), dpi=300)  #adjust overall size if needed
ax = sns.heatmap(corr_matrix, annot=p_value_matrix, fmt='.3f', cmap='coolwarm', 
                  linewidths=.5, annot_kws={"size": 8}, mask=mask, 
                  vmin=-1, vmax=1, cbar_kws={'shrink': 0.5})
plt.title('Kendall Correlation Matrix (w/ P-values)', size=12, weight='bold')

# #plot heatmap of p-values w/ kendal_tau correlations as numbers 
# fig = plt.figure(figsize=(20, 20), dpi=300)  #adjust overall size if needed
# ax = sns.heatmap(p_value_matrix, annot=corr_matrix, fmt='.3f', cmap='coolwarm', 
#                  linewidths=.5, annot_kws={"size": 8}, mask=mask, 
#                  vmin=-1, vmax=1, cbar_kws={'shrink': 0.5})
# plt.title('Kendall Correlation Matrix (w/ P-values)', size=12, weight='bold')

#adjust the aspect ratio of the cells to make them more rectangular
ax.set_aspect("equal")
aspect_ratio = 0.75  #adjust this value to change the cell width
ax.set_aspect(aspect_ratio / ax.get_data_ratio(), adjustable='box')

#increase axes label text size
plt.xticks(fontsize=12, rotation=90)
plt.yticks(fontsize=12)

#increase colorbar text size
cbar = ax.collections[0].colorbar
cbar.ax.tick_params(labelsize=12)
cbar.set_label(label='Kendall Correlation Coefficient', labelpad=7, size=12, weight='bold')

#make significant p-values stand out (bold and italic)
for text in ax.texts:
    p_value = float(text.get_text())
    if p_value < 0.05:
        text.set_weight('bold')
        text.set_style('normal')
        text.set_color('black')
    else:
        text.set_size(6)
        text.set_color('black')
# =============================================================================


# =============================================================================
# Save Correlation Matrix as an Excel Sheet (w/ tau values bolded if sig p-value)
# =============================================================================
from openpyxl import load_workbook
from openpyxl.styles import Font

current_directory = Path.cwd()
absolute_path = Path("local data/")
filename = Path("PALLTER_FullGrid_CorrelationMatrix.xlsx")
savepath = str(current_directory / absolute_path / filename)
bold_threshold = 0.05

corr_df = corr_matrix
pval_df = p_value_matrix

# Mask upper triangle
mask = np.tril(np.ones(corr_df.shape), k=0).astype(bool)
corr_df = corr_df.where(mask)
pval_df = pval_df.where(mask)

with pd.ExcelWriter(savepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    corr_df.to_excel(writer, sheet_name='Correlation')
    pval_df.to_excel(writer, sheet_name='P-Values')

wb = load_workbook(savepath)
ws = wb['Correlation']

# Bold significant tau values; start at row=1, col=1 to skip headers
for row_idx, row in enumerate(corr_df.index, start=2):
    for col_idx, col in enumerate(corr_df.columns, start=2):
        pval = pval_df.loc[row, col]
        if pval < bold_threshold:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(bold=True)

# Bold significant p-values too
ws_pval = wb['P-Values']
for row_idx, row in enumerate(pval_df.index, start=2):
    for col_idx, col in enumerate(pval_df.columns, start=2):
        if pval_df.loc[row, col] < bold_threshold:
            ws_pval.cell(row=row_idx, column=col_idx).font = Font(bold=True)

# Save with formatting
wb.save(savepath)
# =============================================================================
